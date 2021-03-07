import openpyxl

# Read

workbook = openpyxl.load_workbook(r'.\excel\data_read.xlsx')
worksheet = workbook.get_sheet_by_name('Sheet1')
row3 = [item.value for item in list(worksheet.rows)[2]]
print('Row 3 value: ', row3)

col3 = [item.value for item in list(worksheet.columns)[2]]
print("Col 3 value: ", col3)

cell_2_3 = worksheet.cell(row=2,column=3).value
print('Row 2 Col 3 value: ', cell_2_3)

max_row = worksheet.max_row
print('Max row: ', max_row)

# Write

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet['A1'] = 'hi'
workbook.save(r'.\excel\data_write_3.xlsx')